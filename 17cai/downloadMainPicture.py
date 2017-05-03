import requests
import openpyxl
import os

MAIN_PICTURE_FILE_PATH = 'resources/主图信息.xlsx'
HTTP = 'http://'
DOMAIN_NAME = 'http://www.17cai.com'
DIR_PATH = "G:/齐采网/"

os.makedirs(DIR_PATH, exist_ok=True)
workbook = openpyxl.load_workbook(MAIN_PICTURE_FILE_PATH)
main_picture_sheet = workbook.get_active_sheet()
for row in range(2, main_picture_sheet.max_row + 1):
    product_id = str(main_picture_sheet.cell(row=row, column=1).value)
    large_image_url = main_picture_sheet.cell(row=row, column=3).value
    if not large_image_url.startswith(HTTP):
        large_image_url = DOMAIN_NAME + large_image_url
    print(product_id + ":" + large_image_url)

    base_path = os.path.join(DIR_PATH, product_id)
    os.makedirs(base_path, exist_ok=True)
    base_name = os.path.basename(large_image_url)
    picture_path = os.path.join(base_path, base_name)
    picture_file = open(picture_path, 'wb')
    print('Download image path is : ' + picture_path)

    image_stream = requests.get(large_image_url)
    for chunk in image_stream.iter_content(100000):
        picture_file.write(chunk)
