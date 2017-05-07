import threading
import datetime
import requests
import openpyxl
import os

MAIN_PICTURE_FILE_PATH = 'resources/主图信息.xlsx'
HTTP = 'http://'
DOMAIN_NAME = 'http://www.17cai.com'
DIR_PATH = "G:/齐采网/"
INCREMENT_NUMBER = 100
PRODUCT_ID_COLUMN_NUM = 1
LARGE_IMAGE_URL_COLUMN_NUM = 3
CHUNK_SIZE = 100000


def downloadMainPicture(sheet, start, end):
    for row in range(start, end):
        product_id = str(sheet.cell(row=row, column=PRODUCT_ID_COLUMN_NUM).value)
        large_image_url = sheet.cell(row=row, column=LARGE_IMAGE_URL_COLUMN_NUM).value
        if not large_image_url.startswith(HTTP):
            large_image_url = DOMAIN_NAME + large_image_url
        print(product_id + ":" + large_image_url)

        base_path = os.path.join(DIR_PATH, product_id)
        os.makedirs(base_path, exist_ok=True)
        base_name = os.path.basename(large_image_url)
        picture_path = os.path.join(base_path, base_name)
        if os.path.exists(picture_path):
            print('image is exists.')
            continue
        picture_file = open(picture_path, 'wb')
        print('Download image path is : ' + picture_path)

        image_stream = requests.get(large_image_url)
        for chunk in image_stream.iter_content(CHUNK_SIZE):
            picture_file.write(chunk)


print('start time is %s' % datetime.datetime.now())
os.makedirs(DIR_PATH, exist_ok=True)
workbook = openpyxl.load_workbook(MAIN_PICTURE_FILE_PATH)
main_picture_sheet = workbook.get_active_sheet()
downloadThreads = []
for row in range(2, main_picture_sheet.max_row + 1, INCREMENT_NUMBER):
    downloadThread = threading.Thread(target=downloadMainPicture,
                                      args=(main_picture_sheet, row, row + INCREMENT_NUMBER))
    downloadThreads.append(downloadThread)
    downloadThread.start()

for downloadThread in downloadThreads:
    downloadThread.join()

print('end time is %s' % datetime.datetime.now())
print('Done.')
